"""
build_report.py
===============
Excel report builder for risk_engine.

Called from portfolio_demo.py after all risk calculations are done:

    data = run_risk_report(...)          # compute everything, get dict back
    build_excel_report(data, "risk_report.xlsx")   # write the workbook

Entry point
-----------
build_excel_report(d, output_path)
    d           : dict returned by run_risk_report()
    output_path : full path to the .xlsx file to write
"""

import numpy as np
import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colours ────────────────────────────────────────────────────────────────
NAVY        = "1F3864"
DARK_GREY   = "404040"
MID_GREY    = "808080"
LIGHT_GREY  = "F2F2F2"
WHITE       = "FFFFFF"
GREEN_BG    = "E2EFDA"
GREEN_TXT   = "375623"
YELLOW_BG   = "FFEB9C"
YELLOW_TXT  = "9C5700"
RED_BG      = "FFC7CE"
RED_TXT     = "9C0006"
BLUE_INPUT  = "0000FF"
HEADER_FILL = "1F3864"
ALT_ROW     = "EEF2F8"

# ── Style helpers ──────────────────────────────────────────────────────────
def _font(bold=False, size=10, color=DARK_GREY, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _pct_fmt(decimals=2):
    return f'0.{"0"*decimals}%'

def _num_fmt(decimals=3):
    return f'0.{"0"*decimals}'

def header_row(ws, row, cols, labels, widths=None):
    """Write a styled header row."""
    for i, (col, label) in enumerate(zip(cols, labels)):
        c = ws.cell(row=row, column=col, value=label)
        c.font   = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill   = _fill(HEADER_FILL)
        c.alignment = _align("center")
        c.border = _border()
        if widths and i < len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[i]

def section_title(ws, row, col, text, colspan=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font  = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill  = _fill(NAVY)
    c.alignment = _align("left")
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)

def data_cell(ws, row, col, value, fmt=None, bold=False, align="right",
              color=DARK_GREY, fill_color=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=10, color=color)
    c.alignment = _align(align, "center")
    if fmt:
        c.number_format = fmt
    if fill_color:
        c.fill = _fill(fill_color)
    if border:
        c.border = _border()
    return c

def verdict_cell(ws, row, col, text, good=True):
    bg  = GREEN_BG  if good else RED_BG
    txt = GREEN_TXT if good else RED_TXT
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", bold=True, size=10, color=txt)
    c.fill = _fill(bg)
    c.alignment = _align("center", "center")
    c.border = _border()

def zone_cell(ws, row, col, zone, exceptions):
    fills = {"Green": (GREEN_BG, GREEN_TXT),
             "Yellow":(YELLOW_BG, YELLOW_TXT),
             "Red":   (RED_BG,    RED_TXT)}
    bg, txt = fills.get(zone, (LIGHT_GREY, DARK_GREY))
    c = ws.cell(row=row, column=col,
                value=f"{zone}  ({exceptions} exceptions)")
    c.font = Font(name="Arial", bold=True, size=10, color=txt)
    c.fill = _fill(bg)
    c.alignment = _align("center", "center")
    c.border = _border()



# ══════════════════════════════════════════════════════════════════════════
#  SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════

def build_summary(wb, d):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2   # margin

    # ── Title block ────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6
    ws.merge_cells("B2:I2")
    c = ws["B2"]
    c.value     = "PORTFOLIO RISK REPORT"
    c.font      = Font(name="Arial", bold=True, size=16, color=WHITE)
    c.fill      = _fill(NAVY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:I3")
    c = ws["B3"]
    snap = d["snapshot_date"]
    c.value = (f"Snapshot date: {snap.strftime('%d %b %Y')}     "
               f"Report generated: {datetime.today().strftime('%d %b %Y')}     "
               f"Observations: {d['T']}     "
               f"Confidence: 99%     Horizon: 1 day")
    c.font  = Font(name="Arial", size=9, color=MID_GREY, italic=True)
    c.fill  = _fill(LIGHT_GREY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[3].height = 18

    ws.row_dimensions[4].height = 10

    # ── Key metrics ────────────────────────────────────────────────────────
    section_title(ws, 5, 2, "KEY RISK METRICS — 1-DAY 99%", colspan=7)
    ws.row_dimensions[5].height = 20

    header_row(ws, 6, [2,3,4,5,6,7,8],
               ["Measure","Historical VaR","FHS VaR",
                "Historical CVaR","FHS CVaR","Current Vol","Average Vol"],
               widths=[26, 16, 16, 16, 16, 14, 14])

    row = 7
    metrics = [
        ("Portfolio (equity + FX)",
         d["var_hist"], d["var_fhs"], d["cvar_hist"], d["cvar_fhs"],
         d["sigma"][-1], d["sigma"].mean()),
    ]
    for label, vh, vf, ch, cf, sv, av in metrics:
        data_cell(ws, row, 2, label, align="left", border=True, fill_color=WHITE)
        for col, val in zip([3,4,5,6], [vh, vf, ch, cf]):
            data_cell(ws, row, col, val, fmt="0.000%", bold=True,
                      fill_color=WHITE)
        data_cell(ws, row, 7, sv, fmt="0.00%", fill_color=WHITE)
        data_cell(ws, row, 8, av, fmt="0.00%", fill_color=WHITE)
        ws.row_dimensions[row].height = 18
        row += 1

    regime = "ELEVATED" if d["var_fhs"] > d["var_hist"] else "SUBDUED"
    regime_color = YELLOW_BG if regime == "ELEVATED" else GREEN_BG
    regime_txt   = YELLOW_TXT if regime == "ELEVATED" else GREEN_TXT
    ws.merge_cells(f"B{row}:H{row}")
    c = ws.cell(row=row, column=2,
                value=f"Current volatility regime: {regime}  "
                      f"(FHS VaR {'>' if d['var_fhs']>d['var_hist'] else '<'} Historical VaR)")
    c.font = Font(name="Arial", size=9, bold=True, color=regime_txt)
    c.fill = _fill(regime_color)
    c.alignment = _align("left", "center")
    ws.row_dimensions[row].height = 16
    row += 2

    # ── Backtest status ────────────────────────────────────────────────────
    section_title(ws, row, 2, f"BACKTEST STATUS — ROLLING {d['window']}-DAY WINDOW", colspan=7)
    ws.row_dimensions[row].height = 20
    row += 1

    header_row(ws, row, [2,3,4,5,6,7,8],
               ["Test","Result","p-value","Exceptions",
                "Expected","Exception Rate","Detail"])
    ws.row_dimensions[row].height = 18
    row += 1

    bts = d["bts"]

    # Binomial
    data_cell(ws, row, 2, "Binomial (frequency)", align="left", fill_color=WHITE)
    verdict_cell(ws, row, 3, "REJECT" if bts.binomial.reject else "PASS",
                 good=not bts.binomial.reject)
    data_cell(ws, row, 4, bts.binomial.p_value, fmt="0.0000", fill_color=WHITE)
    data_cell(ws, row, 5, bts.binomial.exceptions, fmt="0", fill_color=WHITE)
    data_cell(ws, row, 6, bts.binomial.expected, fmt="0.0", fill_color=WHITE)
    data_cell(ws, row, 7, bts.binomial.actual_rate, fmt="0.00%", fill_color=WHITE)
    data_cell(ws, row, 8, "Too many exceptions → recalibrate" if bts.binomial.reject
              else "Exception count consistent with model",
              align="left", fill_color=WHITE)
    ws.row_dimensions[row].height = 18
    row += 1

    # Christoffersen IND
    data_cell(ws, row, 2, "Christoffersen (independence)", align="left", fill_color=ALT_ROW)
    verdict_cell(ws, row, 3,
                 "REJECT" if bts.christoffersen.reject_independence else "PASS",
                 good=not bts.christoffersen.reject_independence)
    data_cell(ws, row, 4, bts.christoffersen.p_value_independence, fmt="0.0000", fill_color=ALT_ROW)
    data_cell(ws, row, 5, "", fill_color=ALT_ROW)
    data_cell(ws, row, 6, "", fill_color=ALT_ROW)
    data_cell(ws, row, 7, "", fill_color=ALT_ROW)
    data_cell(ws, row, 8, "Exceptions clustering in crises → volatility model too slow"
              if bts.christoffersen.reject_independence
              else "Exceptions independent — no clustering detected",
              align="left", fill_color=ALT_ROW)
    ws.row_dimensions[row].height = 18
    row += 1

    # CVaR exceedance
    if bts.cvar_test is not None and not np.isnan(bts.cvar_test.ratio):
        ratio = bts.cvar_test.ratio
        ratio_flag = ratio > 1.2
        data_cell(ws, row, 2, "CVaR exceedance (tail severity)", align="left", fill_color=WHITE)
        verdict_cell(ws, row, 3,
                     "FLAG" if ratio_flag else "OK",
                     good=not ratio_flag)
        data_cell(ws, row, 4,
                  bts.cvar_test.p_value if not np.isnan(bts.cvar_test.p_value) else "n/a",
                  fmt="0.0000" if not np.isnan(bts.cvar_test.p_value) else "@",
                  fill_color=WHITE)
        data_cell(ws, row, 5, bts.cvar_test.n_exceptions, fmt="0", fill_color=WHITE)
        data_cell(ws, row, 6, "", fill_color=WHITE)
        data_cell(ws, row, 7, ratio, fmt="0.000", fill_color=WHITE)
        data_cell(ws, row, 8,
                  f"Ratio {ratio:.3f} — tail losses exceed CVaR forecast → flag for review"
                  if ratio_flag
                  else f"Ratio {ratio:.3f} — CVaR well calibrated",
                  align="left", fill_color=WHITE)
    else:
        data_cell(ws, row, 2, "CVaR exceedance (tail severity)", align="left", fill_color=WHITE)
        for col in [3,4,5,6,7]:
            data_cell(ws, row, col, "n/a", fill_color=WHITE)
        data_cell(ws, row, 8,
                  f"Insufficient data ({bts.cvar_test.n_exceptions if bts.cvar_test else 0} exceptions) — needs ~10 years",
                  align="left", fill_color=WHITE)
    ws.row_dimensions[row].height = 18
    row += 1

    # Basel
    data_cell(ws, row, 2, "Basel Traffic Light", align="left", fill_color=ALT_ROW)
    zone_cell(ws, row, 3, bts.basel.zone, bts.basel.exceptions)
    data_cell(ws, row, 4, "", fill_color=ALT_ROW)
    data_cell(ws, row, 5, bts.basel.exceptions, fmt="0", fill_color=ALT_ROW)
    data_cell(ws, row, 6, "", fill_color=ALT_ROW)
    data_cell(ws, row, 7, "", fill_color=ALT_ROW)
    data_cell(ws, row, 8,
              f"Capital multiplier: {bts.basel.multiplier:.2f}x  [regulatory — not a statistical test]",
              align="left", fill_color=ALT_ROW)
    ws.row_dimensions[row].height = 18
    row += 2

    # ── Attribution summary ────────────────────────────────────────────────
    section_title(ws, row, 2, "RISK ATTRIBUTION — COMPONENT VaR (COVARIANCE MODEL)", colspan=7)
    ws.row_dimensions[row].height = 20
    row += 1

    header_row(ws, row, [2,3,4,5,6,7,8],
               ["Desk / Book","Security","Weight",
                "Component VaR","% of Total","",""])
    ws.row_dimensions[row].height = 18
    row += 1

    comp_var = pd.Series(d["cov"]["component_var"], index=d["tickers"])
    total_var = d["cov"]["var"]
    alt = False
    for sec, cv in comp_var.sort_values(ascending=False).items():
        desk = d["hierarchy"].loc[sec, "Desk"]
        book = d["hierarchy"].loc[sec, "Book"]
        wt   = d["weights"][d["tickers"].index(sec)]
        fc   = ALT_ROW if alt else WHITE
        data_cell(ws, row, 2, f"{desk} / {book}", align="left", fill_color=fc)
        data_cell(ws, row, 3, sec, align="left", fill_color=fc)
        data_cell(ws, row, 4, wt, fmt="0.0%", fill_color=fc)
        data_cell(ws, row, 5, cv, fmt="+0.000%;-0.000%", fill_color=fc)
        data_cell(ws, row, 6, cv / total_var, fmt="0.0%", fill_color=fc)
        data_cell(ws, row, 7, "", fill_color=fc)
        data_cell(ws, row, 8, "", fill_color=fc)
        ws.row_dimensions[row].height = 16
        alt = not alt
        row += 1

    data_cell(ws, row, 2, "TOTAL", bold=True, align="left", fill_color=LIGHT_GREY)
    data_cell(ws, row, 3, "", fill_color=LIGHT_GREY)
    data_cell(ws, row, 4, sum(d["weights"]), fmt="0.0%", bold=True, fill_color=LIGHT_GREY)
    data_cell(ws, row, 5, total_var, fmt="+0.000%;-0.000%",
              bold=True, fill_color=LIGHT_GREY)
    data_cell(ws, row, 6, 1.0, fmt="0.0%", bold=True, fill_color=LIGHT_GREY)
    data_cell(ws, row, 7, "", fill_color=LIGHT_GREY)
    data_cell(ws, row, 8,
              f"Covariance VaR: {total_var*100:.3f}%  |  "
              f"Historical VaR: {d['var_hist']*100:.3f}%  "
              f"← use Historical for risk decisions",
              align="left", fill_color=LIGHT_GREY,
              color=MID_GREY)
    ws.row_dimensions[row].height = 18
    row += 2

    # ── FX decomposition ──────────────────────────────────────────────────
    section_title(ws, row, 2, "FX RISK DECOMPOSITION — 1-DAY 99%", colspan=7)
    ws.row_dimensions[row].height = 20
    row += 1

    header_row(ws, row, [2,3,4,5,6,7,8],
               ["Component","","VaR","CVaR","","FX add-on",""])
    ws.row_dimensions[row].height = 18
    row += 1

    fx_rows = [
        ("Total (equity + FX)",      d["var_total"],      d["cvar_total"],   WHITE),
        ("Equity (FX frozen)",        d["var_equity"],     d["cvar_equity"],  ALT_ROW),
        ("FX only (equity frozen)",   d["var_fx_only"],    d["cvar_fx_only"], WHITE),
    ]
    for label, v, c, fc in fx_rows:
        data_cell(ws, row, 2, label, align="left", fill_color=fc)
        data_cell(ws, row, 3, "", fill_color=fc)
        data_cell(ws, row, 4, v, fmt="0.000%", fill_color=fc)
        data_cell(ws, row, 5, c, fmt="0.000%", fill_color=fc)
        data_cell(ws, row, 6, "", fill_color=fc)
        data_cell(ws, row, 7, "", fill_color=fc)
        data_cell(ws, row, 8, "", fill_color=fc)
        ws.row_dimensions[row].height = 16
        row += 1

    addon = d["var_total"] - d["var_equity"]
    addon_color = YELLOW_BG if addon > 0 else GREEN_BG
    addon_txt   = YELLOW_TXT if addon > 0 else GREEN_TXT
    ws.merge_cells(f"B{row}:H{row}")
    c = ws.cell(row=row, column=2,
                value=f"FX add-on (Total − Equity):  {addon*100:+.3f}%  "
                      f"({'FX increases portfolio risk' if addon>0 else 'FX diversifies portfolio risk'})")
    c.font = Font(name="Arial", size=9, bold=True, color=addon_txt)
    c.fill = _fill(addon_color)
    c.alignment = _align("left", "center")
    ws.row_dimensions[row].height = 16

    # freeze top rows
    ws.freeze_panes = "B7"


def build_risk_measures(wb, d):
    ws = wb.create_sheet("Risk Measures")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.row_dimensions[1].height = 6
    section_title(ws, 2, 2, "RISK MEASURES — 1-DAY 99% CONFIDENCE", colspan=5)
    ws.row_dimensions[2].height = 20

    header_row(ws, 3, [2,3,4,5,6],
               ["Measure","Value (daily)","Value (annual.)","vs. Hist. VaR","Note"],
               widths=[28, 16, 16, 14, 42])
    ws.row_dimensions[3].height = 18

    sigma = d["sigma"]
    rows = [
        ("EWMA Volatility — current",  sigma[-1],     sigma[-1]*np.sqrt(252),   None,
         "Most recent conditional vol estimate (λ=0.94)"),
        ("EWMA Volatility — average",  sigma.mean(),  sigma.mean()*np.sqrt(252), None,
         "Full-sample average of conditional vol"),
        (None, None, None, None, None),
        ("Historical VaR (primary)",   d["var_hist"], d["var_hist"]*np.sqrt(252), 1.0,
         "Empirical 1st percentile — no distribution assumption"),
        ("FHS VaR",                    d["var_fhs"],  d["var_fhs"]*np.sqrt(252),
         d["var_fhs"]/d["var_hist"],
         "EWMA-standardised, rescaled by current vol — regime-aware"),
        (None, None, None, None, None),
        ("Historical CVaR",            d["cvar_hist"],d["cvar_hist"]*np.sqrt(252),
         d["cvar_hist"]/d["var_hist"],
         "Mean loss beyond VaR — coherent measure (Basel IV / FRTB)"),
        ("FHS CVaR",                   d["cvar_fhs"], d["cvar_fhs"]*np.sqrt(252),
         d["cvar_fhs"]/d["var_hist"],
         "EWMA-filtered CVaR — reflects current vol regime"),
    ]

    alt = False
    for r_idx, row_data in enumerate(rows):
        label, daily, annual, ratio, note = row_data
        excel_row = 4 + r_idx
        if label is None:
            ws.row_dimensions[excel_row].height = 8
            continue
        fc = ALT_ROW if alt else WHITE
        data_cell(ws, excel_row, 2, label, align="left", bold="VaR" in label and "primary" in (note or ""), fill_color=fc)
        if daily is not None:
            data_cell(ws, excel_row, 3, daily, fmt="0.000%", fill_color=fc)
        if annual is not None:
            data_cell(ws, excel_row, 4, annual, fmt="0.00%", fill_color=fc)
        if ratio is not None:
            data_cell(ws, excel_row, 5, ratio, fmt="0.00x", fill_color=fc)
        else:
            data_cell(ws, excel_row, 5, "—", align="center", fill_color=fc)
        data_cell(ws, excel_row, 6, note or "", align="left", color=MID_GREY, fill_color=fc)
        ws.row_dimensions[excel_row].height = 18
        alt = not alt

    regime = "ELEVATED" if d["var_fhs"] > d["var_hist"] else "SUBDUED"
    note_row = 4 + len(rows) + 1
    ws.merge_cells(f"B{note_row}:F{note_row}")
    c = ws.cell(row=note_row, column=2,
                value=f"Current vol regime: {regime}. "
                      f"FHS VaR {'>' if d['var_fhs']>d['var_hist'] else '<'} Historical VaR indicates "
                      f"{'current stress exceeds historical average' if d['var_fhs']>d['var_hist'] else 'current conditions are calmer than historical average'}.")
    c.font = Font(name="Arial", size=9, italic=True, color=MID_GREY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[note_row].height = 16

    ws.merge_cells(f"B{note_row+1}:F{note_row+1}")
    c = ws.cell(row=note_row+1, column=2,
                value="Annualised figures use √252 scaling. Use daily figures for 1-day risk decisions.")
    c.font = Font(name="Arial", size=9, italic=True, color=MID_GREY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[note_row+1].height = 14


def build_attribution(wb, d):
    ws = wb.create_sheet("Attribution")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.row_dimensions[1].height = 6
    section_title(ws, 2, 2, "RISK ATTRIBUTION — COMPONENT VaR (COVARIANCE MODEL, NORMAL)", colspan=6)
    ws.row_dimensions[2].height = 20

    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = (f"Ledoit-Wolf shrinkage δ = {d['cov']['shrinkage']:.2f}   |   "
               f"Covariance VaR = {d['cov']['var']*100:.3f}%   |   "
               f"Historical VaR = {d['var_hist']*100:.3f}%   ← use Historical for risk decisions")
    c.font = Font(name="Arial", size=9, italic=True, color=MID_GREY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[3].height = 16

    # --- By Security ---
    section_title(ws, 5, 2, "BY SECURITY", colspan=6)
    header_row(ws, 6, [2,3,4,5,6,7],
               ["Security","Desk","Book","Weight","Component VaR","% of Total"],
               widths=[22, 14, 14, 10, 16, 12])

    comp_var  = pd.Series(d["cov"]["component_var"], index=d["tickers"])
    total_var = d["cov"]["var"]
    alt = False
    r = 7
    for sec, cv in comp_var.sort_values(ascending=False).items():
        fc = ALT_ROW if alt else WHITE
        wt = d["weights"][d["tickers"].index(sec)]
        data_cell(ws, r, 2, sec,   align="left", fill_color=fc)
        data_cell(ws, r, 3, d["hierarchy"].loc[sec,"Desk"], align="left", fill_color=fc)
        data_cell(ws, r, 4, d["hierarchy"].loc[sec,"Book"], align="left", fill_color=fc)
        data_cell(ws, r, 5, wt, fmt="0.0%", fill_color=fc)
        data_cell(ws, r, 6, cv, fmt="+0.000%;-0.000%", fill_color=fc)
        data_cell(ws, r, 7, cv/total_var, fmt="0.0%", fill_color=fc)
        ws.row_dimensions[r].height = 16
        alt = not alt; r += 1

    # Total row
    data_cell(ws, r, 2, "TOTAL", bold=True, align="left", fill_color=LIGHT_GREY)
    data_cell(ws, r, 3, "", fill_color=LIGHT_GREY)
    data_cell(ws, r, 4, "", fill_color=LIGHT_GREY)
    data_cell(ws, r, 5, sum(d["weights"]), fmt="0.0%", bold=True, fill_color=LIGHT_GREY)
    data_cell(ws, r, 6, total_var, fmt="+0.000%;-0.000%", bold=True, fill_color=LIGHT_GREY)
    data_cell(ws, r, 7, 1.0, fmt="0.0%", bold=True, fill_color=LIGHT_GREY)
    ws.row_dimensions[r].height = 18
    r += 2

    # --- By Book ---
    section_title(ws, r, 2, "BY BOOK", colspan=6)
    ws.row_dimensions[r].height = 20; r += 1
    header_row(ws, r, [2,3,4,5,6,7],
               ["Book","","","","Component VaR","% of Total"])
    ws.row_dimensions[r].height = 18; r += 1

    book_var = comp_var.groupby(d["hierarchy"]["Book"]).sum().sort_values(ascending=False)
    alt = False
    for book, cv in book_var.items():
        fc = ALT_ROW if alt else WHITE
        data_cell(ws, r, 2, book, align="left", fill_color=fc)
        for col in [3,4,5]: data_cell(ws, r, col, "", fill_color=fc)
        data_cell(ws, r, 6, cv, fmt="+0.000%;-0.000%", fill_color=fc)
        data_cell(ws, r, 7, cv/total_var, fmt="0.0%", fill_color=fc)
        ws.row_dimensions[r].height = 16; alt = not alt; r += 1

    data_cell(ws, r, 2, "TOTAL", bold=True, align="left", fill_color=LIGHT_GREY)
    for col in [3,4,5]: data_cell(ws, r, col, "", fill_color=LIGHT_GREY)
    data_cell(ws, r, 6, total_var, fmt="+0.000%;-0.000%", bold=True, fill_color=LIGHT_GREY)
    data_cell(ws, r, 7, 1.0, fmt="0.0%", bold=True, fill_color=LIGHT_GREY)
    ws.row_dimensions[r].height = 18; r += 2

    # --- By Desk ---
    section_title(ws, r, 2, "BY DESK", colspan=6)
    ws.row_dimensions[r].height = 20; r += 1
    header_row(ws, r, [2,3,4,5,6,7],
               ["Desk","","","","Component VaR","% of Total"])
    ws.row_dimensions[r].height = 18; r += 1

    desk_var = comp_var.groupby(d["hierarchy"]["Desk"]).sum().sort_values(ascending=False)
    alt = False
    for desk, cv in desk_var.items():
        fc = ALT_ROW if alt else WHITE
        data_cell(ws, r, 2, desk, align="left", fill_color=fc)
        for col in [3,4,5]: data_cell(ws, r, col, "", fill_color=fc)
        data_cell(ws, r, 6, cv, fmt="+0.000%;-0.000%", fill_color=fc)
        data_cell(ws, r, 7, cv/total_var, fmt="0.0%", fill_color=fc)
        ws.row_dimensions[r].height = 16; alt = not alt; r += 1

    data_cell(ws, r, 2, "TOTAL", bold=True, align="left", fill_color=LIGHT_GREY)
    for col in [3,4,5]: data_cell(ws, r, col, "", fill_color=LIGHT_GREY)
    data_cell(ws, r, 6, total_var, fmt="+0.000%;-0.000%", bold=True, fill_color=LIGHT_GREY)
    data_cell(ws, r, 7, 1.0, fmt="0.0%", bold=True, fill_color=LIGHT_GREY)
    ws.row_dimensions[r].height = 18


def build_backtest(wb, d):
    ws = wb.create_sheet("Backtest")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    bts = d["bts"]
    ws.row_dimensions[1].height = 6
    section_title(ws, 2, 2, f"MODEL VALIDATION — ROLLING {d['window']}-DAY BACKTEST @ 99%", colspan=7)
    ws.row_dimensions[2].height = 20

    ws.merge_cells("B3:H3")
    c = ws["B3"]
    c.value = ("Four tests answer four distinct questions. "
               "Read them in order: frequency → clustering → severity → regulatory.")
    c.font = Font(name="Arial", size=9, italic=True, color=MID_GREY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[3].height = 16

    header_row(ws, 4, [2,3,4,5,6,7,8],
               ["Test","Statistic","Value","Threshold","Result","","Plain English verdict"],
               widths=[26, 20, 12, 12, 10, 2, 46])
    ws.row_dimensions[4].height = 18

    r = 5

    def test_block(label, stat_name, stat_val, threshold, result_text, verdict_good, explanation, fc):
        nonlocal r
        data_cell(ws, r, 2, label,       align="left",   bold=True,  fill_color=fc)
        data_cell(ws, r, 3, stat_name,   align="left",   fill_color=fc)
        if isinstance(stat_val, float) and not np.isnan(stat_val):
            data_cell(ws, r, 4, stat_val, fmt="0.0000", fill_color=fc)
        else:
            data_cell(ws, r, 4, stat_val if stat_val else "—", align="center", fill_color=fc)
        data_cell(ws, r, 5, threshold,   align="center", fill_color=fc)
        verdict_cell(ws, r, 6, result_text, good=verdict_good)
        data_cell(ws, r, 7, "",          fill_color=fc)
        data_cell(ws, r, 8, explanation, align="left", color=MID_GREY, fill_color=fc)
        ws.row_dimensions[r].height = 18
        r += 1

    # 1. Binomial
    section_title(ws, r, 2, "1. BINOMIAL TEST — Frequency of breaches", colspan=7)
    ws.row_dimensions[r].height = 18; r += 1

    test_block("Observed exception rate",
               f"p-value  P(X ≥ {bts.binomial.exceptions})",
               bts.binomial.p_value, "< 0.05",
               "REJECT" if bts.binomial.reject else "PASS",
               not bts.binomial.reject,
               f"{bts.binomial.exceptions} exceptions observed vs {bts.binomial.expected:.1f} expected "
               f"({bts.binomial.actual_rate:.2%} vs {bts.binomial.expected_rate:.2%})",
               WHITE)
    ws.row_dimensions[r].height = 6; r += 1

    # 2. Christoffersen
    section_title(ws, r, 2, "2. CHRISTOFFERSEN TEST — Independence / clustering", colspan=7)
    ws.row_dimensions[r].height = 18; r += 1

    test_block("Unconditional coverage (LR_UC)",
               "p-value  χ²(1)",
               bts.christoffersen.p_value_unconditional, "< 0.05",
               "REJECT" if bts.christoffersen.p_value_unconditional < 0.05 else "PASS",
               bts.christoffersen.p_value_unconditional >= 0.05,
               "Is the total breach count correct? (Same question as Binomial)",
               WHITE)
    test_block("Independence (LR_IND)",
               "p-value  χ²(1)",
               bts.christoffersen.p_value_independence, "< 0.05",
               "REJECT" if bts.christoffersen.reject_independence else "PASS",
               not bts.christoffersen.reject_independence,
               "Are breaches independent? Small p → clustering detected → vol model too slow",
               ALT_ROW)
    test_block("Joint conditional coverage (LR_CC)",
               "p-value  χ²(2)",
               bts.christoffersen.p_value_joint, "< 0.05",
               "REJECT" if bts.christoffersen.reject_joint else "PASS",
               not bts.christoffersen.reject_joint,
               "Both frequency and independence at once. Fails if either component fails.",
               WHITE)
    ws.row_dimensions[r].height = 6; r += 1

    # 3. CVaR exceedance
    section_title(ws, r, 2, "3. CVaR EXCEEDANCE TEST — Tail severity on breach days", colspan=7)
    ws.row_dimensions[r].height = 18; r += 1

    if bts.cvar_test is not None:
        ct = bts.cvar_test
        ratio_flag = not np.isnan(ct.ratio) and ct.ratio > 1.2
        test_block("CVaR exceedance ratio",
                   "ratio = mean(actual) / mean(CVaR)",
                   ct.ratio if not np.isnan(ct.ratio) else None,
                   "> 1.2",
                   "FLAG" if ratio_flag else ("OK" if not np.isnan(ct.ratio) else "n/a"),
                   not ratio_flag,
                   (f"Based on {ct.n_exceptions} exceptions. "
                    + ("p-value: " + f"{ct.p_value:.4f}" if not np.isnan(ct.p_value) else
                       "Insufficient data for t-test (< 3 exceptions) — ratio is qualitative only.") +
                    "  NOTE: test has low power with <3 years of data."),
                   ALT_ROW)
    ws.row_dimensions[r].height = 6; r += 1

    # 4. Basel
    section_title(ws, r, 2, "4. BASEL TRAFFIC LIGHT — Regulatory capital", colspan=7)
    ws.row_dimensions[r].height = 18; r += 1

    zone_colors = {"Green":(GREEN_BG, GREEN_TXT),
                   "Yellow":(YELLOW_BG, YELLOW_TXT),
                   "Red":(RED_BG, RED_TXT)}
    bg, txt = zone_colors.get(bts.basel.zone, (LIGHT_GREY, DARK_GREY))
    data_cell(ws, r, 2, "Zone",           align="left",   bold=True, fill_color=bg, color=txt)
    data_cell(ws, r, 3, "Exceptions / 250", align="left", fill_color=bg, color=txt)
    data_cell(ws, r, 4, bts.basel.exceptions, fmt="0", fill_color=bg, color=txt)
    data_cell(ws, r, 5, "≤ 4 = Green",   align="center", fill_color=bg, color=txt)
    c = ws.cell(row=r, column=6, value=bts.basel.zone)
    c.font = Font(name="Arial", bold=True, size=10, color=txt)
    c.fill = _fill(bg); c.alignment = _align("center"); c.border = _border()
    data_cell(ws, r, 7, "", fill_color=bg)
    data_cell(ws, r, 8,
              f"Capital multiplier: {bts.basel.multiplier:.2f}x  |  "
              "Regulatory output only — not a statistical test.",
              align="left", color=MID_GREY, fill_color=bg)
    ws.row_dimensions[r].height = 18; r += 2

    # Decision tree
    section_title(ws, r, 2, "DECISION FRAMEWORK", colspan=7)
    ws.row_dimensions[r].height = 18; r += 1

    decisions = [
        ("Binomial REJECT",         "Recalibrate the model immediately — exception rate is wrong"),
        ("Christoffersen IND REJECT","Volatility model too slow — consider lower λ or shorter window"),
        ("Basel Yellow / Red",      "Regulatory conversation required regardless of statistical results"),
        ("CVaR ratio > 1.2",        "Flag for senior review — do not act on fewer than 3 years of data"),
        ("All pass",                "Model performing as expected — document results and move on"),
    ]
    alt = False
    for trigger, action in decisions:
        fc = ALT_ROW if alt else WHITE
        data_cell(ws, r, 2, trigger, align="left", bold=True, fill_color=fc)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        data_cell(ws, r, 3, action, align="left", fill_color=fc)
        ws.row_dimensions[r].height = 16; alt = not alt; r += 1


def build_fx(wb, d):
    ws = wb.create_sheet("FX Decomposition")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.row_dimensions[1].height = 6
    section_title(ws, 2, 2, "FX RISK DECOMPOSITION — 1-DAY 99%", colspan=5)
    ws.row_dimensions[2].height = 20

    ws.merge_cells("B3:F3")
    c = ws["B3"]
    c.value = ("r_total = r_equity + r_FX  (exact log decomposition).  "
               "FX add-on = VaR_total − VaR_equity.  "
               "Negative add-on means FX diversifies the portfolio.")
    c.font = Font(name="Arial", size=9, italic=True, color=MID_GREY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[3].height = 16

    header_row(ws, 5, [2,3,4,5,6],
               ["Component","VaR (1-day 99%)","CVaR (1-day 99%)","",""],
               widths=[28, 18, 18, 14, 14])
    ws.row_dimensions[5].height = 18

    fx_rows = [
        ("Total (equity + FX)",     d["var_total"],   d["cvar_total"],   WHITE),
        ("Equity (FX frozen)",       d["var_equity"],  d["cvar_equity"],  ALT_ROW),
        ("FX only (equity frozen)",  d["var_fx_only"], d["cvar_fx_only"], WHITE),
    ]
    r = 6
    for label, v, c_val, fc in fx_rows:
        data_cell(ws, r, 2, label, align="left", fill_color=fc)
        data_cell(ws, r, 3, v, fmt="0.000%", fill_color=fc)
        data_cell(ws, r, 4, c_val, fmt="0.000%", fill_color=fc)
        data_cell(ws, r, 5, "", fill_color=fc)
        data_cell(ws, r, 6, "", fill_color=fc)
        ws.row_dimensions[r].height = 18; r += 1

    addon = d["var_total"] - d["var_equity"]
    addon_color = YELLOW_BG if addon > 0 else GREEN_BG
    addon_txt   = YELLOW_TXT if addon > 0 else GREEN_TXT
    ws.merge_cells(f"B{r}:F{r}")
    c = ws.cell(row=r, column=2,
                value=f"FX add-on: {addon*100:+.3f}%   "
                      f"({'FX exposure increases portfolio risk' if addon>0 else 'FX exposure diversifies portfolio risk'})")
    c.font = Font(name="Arial", bold=True, size=10, color=addon_txt)
    c.fill = _fill(addon_color)
    c.alignment = _align("left", "center")
    ws.row_dimensions[r].height = 20


def build_returns(wb, d):
    ws = wb.create_sheet("Returns")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.row_dimensions[1].height = 6
    section_title(ws, 2, 2, "DAILY LOG-RETURNS AND ROLLING VaR", colspan=4)
    ws.row_dimensions[2].height = 20

    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = ("Daily log-returns: r(t) = log(NAV(t) / NAV(t−1)).  "
               "Rolling VaR uses a 250-day lookback, out-of-sample (no forward-looking bias).")
    c.font = Font(name="Arial", size=9, italic=True, color=MID_GREY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[3].height = 16

    header_row(ws, 5, [2,3,4,5],
               ["Date","Portfolio Return","Rolling VaR (99%)","Rolling CVaR (99%)"],
               widths=[14, 20, 20, 20])
    ws.row_dimensions[5].height = 18

    returns     = d["returns"]
    var_series  = d["var_series"]
    cvar_series = d["cvar_series"]
    dates       = d["return_dates"]

    for i, (dt, ret) in enumerate(zip(dates, returns)):
        r = 6 + i
        ws.row_dimensions[r].height = 14
        c_date = ws.cell(row=r, column=2, value=dt.strftime("%Y-%m-%d") if hasattr(dt,"strftime") else str(dt))
        c_date.font = Font(name="Arial", size=9, color=DARK_GREY)
        c_date.alignment = _align("center")

        ret_val = float(ret)
        c_ret = ws.cell(row=r, column=3, value=ret_val)
        c_ret.number_format = "0.000%"
        c_ret.font = Font(name="Arial", size=9,
                          color=RED_TXT if ret_val < 0 else GREEN_TXT)
        c_ret.alignment = _align("right")

        v = var_series[i] if not np.isnan(var_series[i]) else None
        c_var = ws.cell(row=r, column=4, value=v)
        if v is not None:
            c_var.number_format = "0.000%"
            c_var.font = Font(name="Arial", size=9, color=DARK_GREY)
        else:
            c_var.value = "—"
            c_var.font = Font(name="Arial", size=9, color=MID_GREY)
        c_var.alignment = _align("right")

        cv = cvar_series[i] if not np.isnan(cvar_series[i]) else None
        c_cv = ws.cell(row=r, column=5, value=cv)
        if cv is not None:
            c_cv.number_format = "0.000%"
            c_cv.font = Font(name="Arial", size=9, color=DARK_GREY)
        else:
            c_cv.value = "—"
            c_cv.font = Font(name="Arial", size=9, color=MID_GREY)
        c_cv.alignment = _align("right")

    ws.freeze_panes = "B6"


def build_methodology(wb, d):
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80

    ws.row_dimensions[1].height = 6
    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "METHODOLOGY — risk_engine"
    c.font  = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill  = _fill(NAVY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = ("All risk measures use daily log-returns: r(t) = log(NAV(t) / NAV(t−1)).  "
               "Primary confidence level: 99%.  Horizon: 1 day.")
    c.font = Font(name="Arial", size=9, italic=True, color=MID_GREY)
    c.alignment = _align("left", "center")
    ws.row_dimensions[3].height = 16

    sections = [
        ("VOLATILITY", [
            ("EWMA",
             "Exponentially Weighted Moving Average with λ = 0.94 (RiskMetrics convention). "
             "Recursion: σ²_t = λ·σ²_{t-1} + (1−λ)·r²_{t-1}. "
             "Effective memory ≈ 17 days. No distribution assumption. "
             "Known limitations: λ is a convention not an estimate; no mean reversion; "
             "symmetric response to positive and negative shocks."),
        ]),
        ("RISK MEASURES", [
            ("Historical VaR  (PRIMARY)",
             "Negative 1st empirical percentile of the return distribution. "
             "No distribution assumption. Directly uses the actual tail of the observed data. "
             "This is the primary risk number. Use for all risk decisions and limit comparisons."),
            ("FHS VaR",
             "Filtered Historical Simulation. Standardises returns by EWMA vol, takes the "
             "empirical quantile of the residuals, rescales by current vol. "
             "Combines distribution-free tail shape with current volatility regime. "
             "When FHS > Historical, current vol is elevated above the historical average."),
            ("CVaR / Expected Shortfall",
             "Mean loss on the days that exceed the VaR threshold. "
             "CVaR ≥ VaR always. Coherent risk measure mandated under Basel IV / FRTB at 97.5%. "
             "The gap between CVaR and VaR indicates tail heaviness beyond the threshold."),
            ("FHS CVaR",
             "Same standardisation and rescaling as FHS VaR, using tail mean instead of quantile. "
             "Regime-aware: reflects current vol via σ_T."),
        ]),
        ("RISK ATTRIBUTION", [
            ("Component VaR",
             "Decomposes portfolio VaR into per-security contributions via the covariance matrix. "
             "Component VaR_i = w_i × (∂VaR/∂w_i). "
             "Sum of component VaRs = portfolio VaR exactly (Euler homogeneity property). "
             "Uses Normal assumption scoped to decomposition only — not to risk measurement."),
            ("Ledoit-Wolf Shrinkage",
             "Replaces noisy sample covariance with a blend of sample covariance and scaled identity. "
             "Σ* = (1−δ)·S + δ·μ·I where δ is chosen analytically to minimise estimation error. "
             "Meaningful when T/N < 20. δ = 0 means pure sample; δ = 1 means full shrinkage."),
        ]),
        ("MODEL VALIDATION — BACKTESTING", [
            ("Binomial Test",
             "Exact test of whether the exception frequency matches the model's confidence level. "
             "X ~ Binomial(T, 1%). p-value = P(X ≥ observed | H0). "
             "One-tailed upper test: flags too many exceptions only. "
             "No asymptotic approximation — exact for any sample size."),
            ("Christoffersen (1998)",
             "Tests whether exceptions are independent of each other or cluster in crisis periods. "
             "Counts four transition types: calm→calm (n00), calm→breach (n01), "
             "breach→calm (n10), breach→breach (n11). "
             "LR_IND tests H0: P(breach|breach yesterday) = P(breach|no breach yesterday). "
             "LR_CC = LR_UC + LR_IND is the joint test. All statistics are asymptotically chi-squared."),
            ("CVaR Exceedance",
             "On breach days only: tests H0: mean(actual loss − CVaR forecast) = 0 via t-test. "
             "ratio = mean(actual loss) / mean(CVaR forecast). "
             "IMPORTANT: at 99% VaR you get 2–3 breach days per year. "
             "The test needs ~10 years to have meaningful statistical power. "
             "Treat the ratio as a qualitative signal in the short run."),
            ("Basel Traffic Light",
             "Regulatory output — not a statistical test. "
             "Exception count over 250 days → Green (0-4) / Yellow (5-9) / Red (10+). "
             "Thresholds are regulatory conventions from Basel (1996), not statistically derived. "
             "Capital multiplier: Green = 3.00x, Yellow = 3.40–3.85x, Red = 4.00x."),
        ]),
        ("FX DECOMPOSITION", [
            ("Log decomposition",
             "r_total = r_equity + r_FX (exact in log space — approximate in arithmetic returns). "
             "Equity returns computed by freezing FX at the positions snapshot date. "
             "FX add-on = VaR_total − VaR_equity. "
             "Negative add-on = FX exposure is diversifying. "
             "Positive add-on = FX exposure is additive to equity risk."),
        ]),
        ("REFERENCES", [
            ("RiskMetrics (1996)",
             "J.P. Morgan / Reuters Technical Document. EWMA volatility and FHS methodology."),
            ("Ledoit & Wolf (2004)",
             "'Honey, I Shrunk the Sample Covariance Matrix'. Journal of Portfolio Management."),
            ("Christoffersen (1998)",
             "'Evaluating Interval Forecasts'. International Economic Review."),
            ("McNeil & Frey (2000)",
             "'Estimation of Tail-Related Risk Measures for Heteroscedastic Financial Time Series'. "
             "Journal of Empirical Finance. CVaR exceedance test."),
            ("Basel Committee (1996)",
             "'Supervisory Framework for the Use of Backtesting in Conjunction with the "
             "Internal Models Approach to Market Risk Capital Requirements.'"),
        ]),
    ]

    r = 5
    for section_name, items in sections:
        ws.row_dimensions[r].height = 6; r += 1
        section_title(ws, r, 2, section_name, colspan=2)
        ws.row_dimensions[r].height = 18; r += 1

        alt = False
        for term, desc in items:
            fc = ALT_ROW if alt else WHITE
            c_term = ws.cell(row=r, column=2, value=term)
            c_term.font = Font(name="Arial", bold=True, size=9, color=DARK_GREY)
            c_term.fill = _fill(fc)
            c_term.alignment = _align("left", "top", wrap=True)
            c_term.border = _border()

            c_desc = ws.cell(row=r, column=3, value=desc)
            c_desc.font = Font(name="Arial", size=9, color=DARK_GREY)
            c_desc.fill = _fill(fc)
            c_desc.alignment = _align("left", "top", wrap=True)
            c_desc.border = _border()

            # auto-height: ~14pt per 80 chars
            lines = max(1, len(desc) // 78 + 1)
            ws.row_dimensions[r].height = max(18, lines * 14)
            alt = not alt; r += 1


# ══════════════════════════════════════════════════════════════════════════
#  PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════

def build_excel_report(d: dict, output_path: str) -> None:
    """
    Build the full Excel workbook from a risk data dict and save it.

    Parameters
    ----------
    d           : dict returned by portfolio_demo.run_risk_report()
    output_path : full path to the .xlsx file to write (created or overwritten)
    """
    wb = Workbook()

    build_summary(wb, d)
    build_risk_measures(wb, d)
    build_attribution(wb, d)
    build_backtest(wb, d)
    build_fx(wb, d)
    build_returns(wb, d)
    build_methodology(wb, d)

    wb.save(output_path)
